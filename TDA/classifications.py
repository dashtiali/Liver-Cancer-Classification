# Experiment with tda features
# Patient classification via soft voting

# ## Settings:
# 1. Whole Liver
# 2. Largest 15 slices 
# 3. Slice-wise (images labeled with patient ID and slice number)
# 4. Cubical complex -> barcodes (dim 0,1)
# 
# ## Feature Extraction
# 1. Betti Curve
# 2. Entropy Summary
# 3. Pers Landscape
# 4. Pers Stats
# 5. Pers Tropical Coordinates


import os

import numpy as np
import pandas as pd
import shutup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pycaret.classification import *
from sklearn.base import BaseEstimator, TransformerMixin
from catboost import CatBoostClassifier
from lightgbm import LGBMClassifier
from sklearn.ensemble import (AdaBoostClassifier, ExtraTreesClassifier,
                              GradientBoostingClassifier,
                              RandomForestClassifier, StackingClassifier)
from sklearn.linear_model import LassoCV, LogisticRegression, SGDClassifier, RidgeClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.calibration import CalibratedClassifierCV
from sklearn.metrics import (accuracy_score, f1_score, precision_score,
                             recall_score, roc_auc_score, confusion_matrix)
from sklearn.model_selection import StratifiedGroupKFold, train_test_split
from sklearn.preprocessing import StandardScaler
from tqdm import tqdm
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib
# Set the matplotlib backend to 'Agg'
matplotlib.use('Agg')

# Helper functions

def save_confusion_matrix(y_true, y_pred, title, class_names, file_name):
    cm = confusion_matrix(y_true, y_pred)
    plt.figure(figsize=(6, 6))
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', cbar=False, 
                xticklabels=class_names, 
                yticklabels=class_names)
    plt.xlabel('Predicted Label')
    plt.ylabel('True Label')
    plt.title(title)
    plt.savefig(file_name)
    plt.close()

def append_df_to_excel(file_path, sheet_name, model_name, df):
    # Create an empty Excel file if it doesn't exist
    if not os.path.exists(file_path):
        # Create an empty DataFrame to write an empty file
        pd.DataFrame().to_excel(file_path)

    # Load the existing Excel file
    book = load_workbook(file_path)
    
    # Create the new sheet if it doesn't exist
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
        
    # Access the sheet
    sheet = book[sheet_name]

    # Find the last row in the sheet
    last_row = sheet.max_row

    # Write the model name in the first column of the next row
    sheet.cell(row=last_row + 1, column=1, value=model_name)
    
    # Convert DataFrame to a list of rows
    rows = dataframe_to_rows(df, index=False, header=True)
    
    # Write DataFrame to the sheet starting from the second column and the next row
    for i, row in enumerate(rows, start=last_row + 2):
        for j, value in enumerate(row, start=2):  # Start writing from the second column
            sheet.cell(row=i, column=j, value=value)
    
    # Write a double new line
    sheet.cell(row=i + 1, column=1, value='')

    # Save the workbook
    book.save(file_path)

def aggregate_features(featur_dict):
    feature_agg = []

    for i, feature_name in enumerate(featur_dict):
        for j, dim in enumerate(featur_dict[feature_name]):
            feature_df = pd.read_csv(featur_dict[feature_name][dim])

            # Initialize the aggregated feature data frame with the first two columns
            if i == j == 0:
                feature_agg = pd.DataFrame(feature_df[['Patient_ID', 'File_name']])
            
            # Drop the first column
            feature_df_mod = feature_df.drop(columns=['Patient_ID'])

            # Merge DataFrames on 'File_name'
            feature_agg = pd.merge(feature_agg, feature_df_mod, on='File_name')
            
    return feature_agg

def stratified_group_split(data, target_column, group_column, test_size=0.2, random_state=42):
    # Create a DataFrame of unique patient IDs with their class
    unique_patients = data[[group_column, target_column]].drop_duplicates()
    
    integer_portion = int(unique_patients.shape[0] * test_size)
    test_size = integer_portion/unique_patients.shape[0]

    # Perform the split
    train_patients, test_patients = train_test_split(
        unique_patients, 
        test_size=test_size, 
        stratify=unique_patients[target_column], 
        random_state=random_state
    )
    
    # Filter the original DataFrame using train and test patient IDs
    train_df = data[data[group_column].isin(train_patients[group_column])]
    test_df = data[data[group_column].isin(test_patients[group_column])]

    return train_df, test_df

def get_metrics_df(metrics_summary):
    # Convert the metrics summary to a DataFrame
    metrics_df = pd.DataFrame(metrics_summary)
    
    average_metrics = metrics_df.mean(numeric_only=True)
    
    # Convert the Series to a dictionary and add the 'Fold' value
    average_metrics_dict = average_metrics.to_dict()
    average_metrics_dict['Fold'] = 'Average'
    
    # Append the average metrics to the DataFrame
    metrics_df = pd.concat([metrics_df, pd.DataFrame(average_metrics_dict, index=[0])], ignore_index=True)
    
    return metrics_df

def predict_via_soft_voting(X_test, model, rev_class_mapping):
    grouped_patients = X_test.groupby('Patient_ID')

    pred_columns = ['Patient_ID', 'Predicted_label', 'probability']
    patient_predictions = pd.DataFrame(columns=pred_columns)
    
    slice_pred_columns = ['Patient_ID', 'File_name', 'True_label', 'Predicted_label', f'prob_{rev_class_mapping[0]}', f'prob_{rev_class_mapping[1]}']
    slice_predictions = pd.DataFrame(columns=slice_pred_columns)

    for patient_id, slice_ids in grouped_patients.groups.items():
        
        data = X_test.loc[slice_ids].drop(columns=['Class', 'Patient_ID'])
        predicted_prob = model.predict_proba(data)
        
        # Using mean
        predicted_patient_class = np.argmax(predicted_prob.mean(axis=0))
        probability = predicted_prob.mean(axis=0)[1]
        
        # Using median
        # predicted_patient_class = np.argmax(np.median(predicted_prob, axis=0))
        # probability = np.median(predicted_prob, axis=0)[1]
        
        df = pd.DataFrame([[patient_id, predicted_patient_class, probability]], columns=pred_columns)
        
        patient_predictions = pd.concat([patient_predictions, df], ignore_index=True)
        
        predicted_labels = model.predict(data)
        
        slices_df = pd.DataFrame(columns=slice_pred_columns)
        slices_df[slice_pred_columns[0]] = X_test.loc[slice_ids]['Patient_ID']
        slices_df[slice_pred_columns[1]] = slice_ids
        slices_df[slice_pred_columns[2]] = X_test.loc[slice_ids]['Class']
        slices_df[slice_pred_columns[3]] = predicted_labels
        slices_df[slice_pred_columns[4]] = predicted_prob[:, 0]
        slices_df[slice_pred_columns[5]] = predicted_prob[:, 1]
        
        slice_predictions = pd.concat([slice_predictions, slices_df], ignore_index=True)

    get_true_label = lambda x: X_test[X_test['Patient_ID'] == x].iloc[0]['Class']
    patient_predictions['True_labels'] = patient_predictions['Patient_ID'].apply(get_true_label)

    return patient_predictions, slice_predictions


class LassoFeatureSelector(BaseEstimator, TransformerMixin):
    def __init__(self, n_features=10):
        self.n_features = n_features
        self.selected_features = []

    def fit(self, X, y):
        lasso = LassoCV(cv=5, random_state=42)
        lasso.fit(X, y)
        self.selected_features = X.columns[lasso.coef_ != 0].tolist()

        return self

    def transform(self, X):
        return X[self.selected_features]

    def get_feature_names_out(self, input_features=None):
        return self.selected_features


def init_classifier(clf_name, best_model):
    if clf_name == 'Ensemble':
        ensemble_classifiers = []
        for clf in best_model.steps[-1][1].estimators:
            
            ensemble_classifiers.append((type(clf[1]).__name__, clf[1].get_params()))
        
        estimators_list = []
        
        for clf in ensemble_classifiers:
            temp_classifier = globals()[clf[0]](**clf[1])
            temp_clf_name = clf[0]
            
            if temp_clf_name in ['RidgeClassifier', 'SGDClassifier']:
                temp_classifier = CalibratedClassifierCV(temp_classifier)
            
            estimators_list.append((temp_clf_name, temp_classifier))

        classifier = StackingClassifier(
            estimators=estimators_list,
            final_estimator=estimators_list[0][1]
            )
    else:
        hyper_params = best_model.steps[-1][1].get_params()
        classifier = globals()[clf_name](**hyper_params)
        
        if clf_name in ['RidgeClassifier', 'SGDClassifier']:
            classifier = CalibratedClassifierCV(classifier)
    
    return classifier


if __name__ == '__main__':
    shutup.please()
    
    RS = 123

    main_feat_dir = f'extracted_features'
    pycaret_exp_dir = 'initial_results'
    models_dir = f'{pycaret_exp_dir}\\saved_models'

    output_dir = 'final_results'

    if not os.path.isdir(output_dir):
        os.mkdir(output_dir)
    
    plots_dir = f'{output_dir}\\plots'
    
    if not os.path.isdir(plots_dir):
        os.mkdir(plots_dir)

    # Extract feature and trained model names
    best_classifiers = pd.read_csv(f'{pycaret_exp_dir}\\all_fold_results.csv', index_col=0, dtype=str, encoding='latin1')
    best_classifiers = best_classifiers[['Feature', 'Model']]
    best_classifiers['Model'] = best_classifiers['Model'].apply(lambda x: x.replace(' ', ''))
    group_classifiers = best_classifiers.groupby('Feature')

    ph_dims = [0, 1]
    metric_columns = ['Accuracy', 'AUC', 'Recall', 'Precision', 'F1']
    
    empty_df = pd.DataFrame()
    all_features_fold_results_file_path = f'{output_dir}\\all_features_fold_results.xlsx'
    empty_df.to_excel(all_features_fold_results_file_path, index=False)

    avg_fold_result_columns = ['Feature', 'Model'] + [m + ' ± Std.' for m in metric_columns]
    all_features_avg_fold_results_file_path = f'{output_dir}\\all_features_avg_fold_results.csv'
    pd.DataFrame(columns=avg_fold_result_columns).to_csv(all_features_avg_fold_results_file_path, header=True, index=False, encoding="cp1252")

    test_result_columns = ['Feature', 'Model'] + metric_columns
    all_features_test_results_file_path = f'{output_dir}\\all_features_test_results.csv'
    pd.DataFrame(columns=test_result_columns).to_csv(all_features_test_results_file_path, header=True, index=False, encoding="cp1252")
    
    test_patients_pred_results_file_path = f'{output_dir}\\test_patients_pred_results.xlsx'
    empty_df.to_excel(test_patients_pred_results_file_path, index=False)
    
    test_miss_pred_results_file_path = f'{output_dir}\\test_miss_pred_results.xlsx'
    empty_df.to_excel(test_miss_pred_results_file_path, index=False)

    for feature_name, df in tqdm(group_classifiers):
        feature_list = feature_name.replace(' ', '').replace('Persistent', 'Pers').replace('Statistics', 'Stats').split('-')
        print(f'\nProcessing Feature list {feature_name} ....')
        
        short_feature_name = 'Fusion' if len(feature_list) > 1 else feature_name
        
        ########################
        # Load and prepare feature data
        feature_dict_ICC = {feature_name: 
                            {d: f'{main_feat_dir}\\ICC\\feature_matrix_dim_{d}_{feature_name}.csv' for d in ph_dims} 
                            for feature_name in feature_list}

        feature_dict_HCC = {feature_name: 
                            {d: f'{main_feat_dir}\\HCC\\feature_matrix_dim_{d}_{feature_name}.csv' for d in ph_dims} 
                            for feature_name in feature_list}

        ICC_feature = aggregate_features(feature_dict_ICC)
        HCC_feature = aggregate_features(feature_dict_HCC)

        ICC_feature['Class'] = 'ICC'
        HCC_feature['Class'] = 'HCC'

        combined_data = pd.concat([ICC_feature, HCC_feature], ignore_index=True)
        combined_data = combined_data.set_index('File_name')
        train_data, test_data = stratified_group_split(combined_data, target_column='Class', group_column='Patient_ID', test_size=0.2, random_state=RS)
        ########################
        
        ########################
        # Setup group 5-fold
        
        # Convert class labels to numeric
        class_names = combined_data['Class'].unique()
        class_mapping = {name: i for i,name in enumerate(class_names)}
        rev_class_mapping = {v: k for k, v in class_mapping.items()}

        train_data_enc = train_data
        train_data_enc['Class'] = train_data_enc['Class'].map(class_mapping)

        test_data_enc = test_data
        test_data_enc['Class'] = test_data_enc['Class'].map(class_mapping)

        # Initialize StratifiedGroupKFold
        sgkf = StratifiedGroupKFold(n_splits=5, shuffle=True,  random_state=RS)
        
        ########################
        
        selected_features = np.load(f'{pycaret_exp_dir}\\{feature_name}_selected_feature.npy', allow_pickle=True)
        
        for clf_name in df['Model']:
            best_model = load_model(f'{models_dir}\\{feature_name}_tuned_{clf_name}_model', verbose=False)
            classifier = init_classifier(clf_name, best_model)
        
            metrics_summary = []
            # predictions = {'y_true': [], 'y_pred': [], 'Patient_ID': [], 'Fold': []}
            fold_number = 1

            # slice_pred_columns = ['Patient_ID', 'File_name', 'True_label', 'Predicted_label', f'prob_{rev_class_mapping[0]}', f'prob_{rev_class_mapping[1]}']
            # all_slices_predictions = pd.DataFrame(columns=slice_pred_columns)

            # Data for AUC plot
            # patient_true_labels = []
            # patient_predicted_probs = []

            for train_index, val_index in sgkf.split(train_data_enc, train_data_enc['Class'], train_data_enc['Patient_ID']):
                train_df = train_data_enc.iloc[train_index]
                val_df = train_data_enc.iloc[val_index]

                X_train = train_df.drop(columns=['Class', 'Patient_ID'])
                y_train = train_df['Class']
                
                X_val = val_df.drop(columns=['Class', 'Patient_ID'])
                # y_val = val_df['Class']

                # Standardize the features
                scaler = StandardScaler()
                X_train = pd.DataFrame(scaler.fit_transform(X_train), columns=X_train.columns, index=X_train.index)
                # Filter data with selected features
                X_train = X_train[selected_features]

                X_val = pd.DataFrame(scaler.transform(X_val), columns=X_val.columns, index=X_val.index)
                X_val = pd.concat([val_df[['Patient_ID', 'Class']], X_val[selected_features]], axis=1)
                # X_val = pd.concat([val_df[['Patient_ID', 'Class']], X_val], axis=1)

                # Train the classifier
                classifier.fit(X_train, y_train)

                # Predict probabilities
                patient_predictions, slice_predictions = predict_via_soft_voting(X_val, classifier, rev_class_mapping)

                y_val = patient_predictions['True_labels']
                y_pred = patient_predictions['Predicted_label'].astype('Int64')
                y_pred_prob = patient_predictions['probability']
                
                # Calculate metrics
                accuracy = accuracy_score(y_val, y_pred)
                auc_score = roc_auc_score(y_val, y_pred_prob)
                recall = recall_score(y_val, y_pred, pos_label=1)
                precision = precision_score(y_val, y_pred, pos_label=1)
                f1 = f1_score(y_val, y_pred, pos_label=1)
            
                # Store the metrics in a dictionary
                metrics_summary.append({
                    'Fold': fold_number,
                    'Accuracy': accuracy,
                    'AUC': auc_score,
                    'Recall': recall,
                    'Precision': precision,
                    'F1': f1
                })
                
                # all_slices_predictions = pd.concat([all_slices_predictions, slice_predictions], ignore_index=True)
                
                # patient_true_labels.extend(y_val)
                # patient_predicted_probs.extend(y_pred_prob)
                
                # Store predictions, true labels, and probabilities with Patient_ID and fold number
                # predictions['y_true'].extend(y_val)
                # predictions['y_pred'].extend(y_pred)
                # predictions['Patient_ID'].extend(val_df['Patient_ID'])
                # predictions['Fold'].extend([fold_number] * len(y_val))

                fold_number +=1
            
            metrics_df = get_metrics_df(metrics_summary)
            
            append_df_to_excel(all_features_fold_results_file_path, short_feature_name, clf_name, metrics_df)
            
            ###############################
            # Write average fold metrics for current feature and classifier
            formatted_metrics = [short_feature_name, clf_name]
            mean_values = metrics_df.iloc[-1, 1:]
            std_values = metrics_df.iloc[:-1, 1:].std()
            formatted_metrics.extend([f"{mean:.4f} ± {std:.4f}" for mean, std in zip(mean_values, std_values)])

            # Open the CSV file in append mode and write the new record
            with open(all_features_avg_fold_results_file_path, mode='a') as file:
                # Convert the list to a comma-separated string
                new_record_str = ','.join(formatted_metrics) + '\n'
                
                # Write the new record
                file.write(new_record_str)


            ###############################
            # Prediction on the hold out set
            train_df = train_data_enc
            test_df = test_data_enc

            X_train = train_df.drop(columns=['Class', 'Patient_ID'])
            y_train = train_df['Class']

            X_test = test_df.drop(columns=['Class', 'Patient_ID'])

            # Standardize the features
            scaler = StandardScaler()
            X_train = pd.DataFrame(scaler.fit_transform(X_train), columns=X_train.columns, index=X_train.index)
            # Filter data with selected features
            X_train = X_train[selected_features]
            
            X_test = pd.DataFrame(scaler.transform(X_test), columns=X_test.columns, index=X_test.index)
            X_test = pd.concat([test_df[['Patient_ID', 'Class']], X_test[selected_features]], axis=1)

            # Initialize the classifier
            classifier = init_classifier(clf_name, best_model)

            # Train the classifier
            classifier.fit(X_train, y_train)

            # Predict probabilities
            patient_predictions, slice_predictions = predict_via_soft_voting(X_test, classifier, rev_class_mapping)

            y_test = patient_predictions['True_labels']
            y_pred = patient_predictions['Predicted_label'].astype('Int64')
            y_pred_prob = patient_predictions['probability']

            # Calculate metrics
            accuracy = accuracy_score(y_test, y_pred)
            auc_score = roc_auc_score(y_test, y_pred_prob)
            recall = recall_score(y_test, y_pred, pos_label=1)
            precision = precision_score(y_test, y_pred, pos_label=1)
            f1 = f1_score(y_test, y_pred, pos_label=1)

            # Store the metrics in a dictionary
            metrics_summary = [accuracy, auc_score, recall, precision, f1]

            # Write average fold metrics for current feature and classifier
            formatted_metrics = [short_feature_name, clf_name]
            formatted_metrics.extend([f"{m:.4f}" for m in metrics_summary])

            # Open the CSV file in append mode and write the new record
            with open(all_features_test_results_file_path, mode='a') as file:
                # Convert the list to a comma-separated string
                new_record_str = ','.join(formatted_metrics) + '\n'
                
                # Write the new record
                file.write(new_record_str)
            ###############################
            
            patient_predictions['True_labels'] = patient_predictions['True_labels'].map(rev_class_mapping)
            patient_predictions['Predicted_label'] = patient_predictions['Predicted_label'].map(rev_class_mapping)
            
            append_df_to_excel(test_patients_pred_results_file_path, short_feature_name, clf_name, patient_predictions)
            
            miss_predicted_patients = patient_predictions[patient_predictions['Predicted_label'] != patient_predictions['True_labels']]
            append_df_to_excel(test_miss_pred_results_file_path, short_feature_name, clf_name, miss_predicted_patients)
            
            save_confusion_matrix(y_test, y_pred, title=f'{clf_name} Confusion Matrix', class_names=list(class_mapping.keys()), file_name=f'{plots_dir}\\{short_feature_name}_{clf_name}_confusion_matrix.png')
            
        
        print('\n\n')
    